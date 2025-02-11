import openpyxl
from tkinter import Tk, filedialog, simpledialog, messagebox, ttk
import os
import sys

# Suppress command-line window (useful when packaged with PyInstaller)
if getattr(sys, 'frozen', False):
    sys.argv.append('--noconsole')

# Main function
def main():
    # Create the main Tkinter window
    root = Tk()
    root.withdraw()  # Hide the root window

    # Prompt for Excel file
    messagebox.showinfo("Input File", "Please input Excel file")
    file_path = filedialog.askopenfilename(
        title="Select the Excel file",
        filetypes=[("Excel Files", "*.xlsx")],
    )

    if not file_path:
        messagebox.showerror("Error", "No file selected. Exiting program.")
        return

    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active  # Use the active sheet

        # Ask for the quantity column
        quantity_column = simpledialog.askstring(
            "Input", 
            "Enter the column letter for quantities (default is J):", 
            initialvalue="J"
        ).upper()

        if not quantity_column or not quantity_column.isalpha():
            messagebox.showerror("Error", "Invalid column letter entered. Exiting program.")
            return
        
        quantity_col_index = openpyxl.utils.column_index_from_string(quantity_column)
        total_rows = ws.max_row - 1  # Exclude the header row

        # Create a GUI window for the progress bar
        progress_window = Tk()
        progress_window.title("Progress")
        progress_label = ttk.Label(progress_window, text="Processing rows...")
        progress_label.pack(pady=10)
        progress_bar = ttk.Progressbar(progress_window, length=300, mode='determinate')
        progress_bar.pack(pady=10)
        progress_window.update()

        # Display a warning message
        messagebox.showinfo(
            "Notice",
            "This program will attempt to remove all duplicate rows "
            "that differ in any rows but the quantity row. Make sure there "
            "aren't any unique values (e.g., Warehouse ID) that might prevent "
            "the program from removing duplicates correctly."
        )

        # Gather all row data except the quantity column
        data = {}
        for row in range(2, ws.max_row + 1):  # Skip the header row
            row_values = tuple(ws.cell(row=row, column=col).value 
                               for col in range(1, ws.max_column + 1) 
                               if col != quantity_col_index)
            quantity = ws.cell(row=row, column=quantity_col_index).value or 0

            if row_values in data:
                data[row_values]['quantity'] += quantity
            else:
                data[row_values] = {'row': row, 'quantity': quantity}
            
            # Update progress bar
            progress_bar['value'] = (row - 1) / total_rows * 100
            progress_window.update()

        progress_label.config(text="Removing duplicate rows...")
        progress_window.update()

        # Remove duplicate rows and update quantities
        for row in range(ws.max_row, 1, -1):  # Iterate backward for deletion
            row_values = tuple(ws.cell(row=row, column=col).value 
                               for col in range(1, ws.max_column + 1) 
                               if col != quantity_col_index)

            if row_values in data and data[row_values]['row'] != row:
                first_row = data[row_values]['row']
                ws.cell(row=first_row, column=quantity_col_index).value = data[row_values]['quantity']
                ws.delete_rows(row)
            
            # Update progress bar
            progress_bar['value'] = ((ws.max_row - row + 1) / total_rows) * 100
            progress_window.update()

        progress_window.destroy()
        messagebox.showinfo("Success", "Duplicate removal complete.")

        # Save the updated file
        new_file_path = os.path.splitext(file_path)[0] + "_summed.xlsx"
        wb.save(new_file_path)
        messagebox.showinfo("Success", f"File saved as '{new_file_path}'.")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

if __name__ == "__main__":
    main()
